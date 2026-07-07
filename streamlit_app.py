"""
════════════════════════════════════════════════════════════════════════════
STREAMLIT APP: Разделение и объединение Excel через ZIP-архивы
С СОХРАНЕНИЕМ ВИЗУАЛЬНОГО ФОРМАТИРОВАНИЯ (цвета, шрифты, границы)
════════════════════════════════════════════════════════════════════════════

Ключевое отличие от pandas-версии:
    Используем openpyxl напрямую вместо pandas.read_excel/to_excel,
    потому что pandas работает только со ЗНАЧЕНИЯМИ ячеек и стирает
    всё визуальное форматирование (заливку, шрифт, границы).

    openpyxl видит каждую ячейку как объект со свойствами:
        cell.value  — значение
        cell.font   — шрифт (цвет, жирность)
        cell.fill   — заливка (цвет фона)
        cell.border — границы
        cell.alignment — выравнивание
        cell.number_format — формат чисел

    Копируя все эти свойства построчно, мы сохраняем внешний вид
    исходной таблицы и в разбитых частях, и в итоговом объединённом файле.
"""

import io
import zipfile
import logging
from copy import copy
from datetime import datetime

import pandas as pd  # используется ТОЛЬКО для превью в интерфейсе
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ
# ════════════════════════════════════════════════════════════════════════════

MAX_UNCOMPRESSED_SIZE_MB = 200
MAX_UNCOMPRESSED_SIZE_BYTES = MAX_UNCOMPRESSED_SIZE_MB * 1024 * 1024
ALLOWED_EXCEL_EXTENSIONS = (".xlsx", ".xls")

HEADER_ROW = 1  # предполагаем, что первая строка — заголовки колонок


# ════════════════════════════════════════════════════════════════════════════
# БЛОК УТИЛИТ: КОПИРОВАНИЕ СТИЛЕЙ МЕЖДУ ЯЧЕЙКАМИ
# ════════════════════════════════════════════════════════════════════════════

def copy_cell(src_cell, dst_cell) -> None:
    """
    Копирует значение И визуальное оформление одной ячейки в другую.

    ⚡ Копируются:
        - значение
        - шрифт (цвет текста, жирность, курсив, размер)
        - заливка (цвет фона ячейки — то, что просили сохранить)
        - границы
        - выравнивание
        - формат чисел (даты, проценты, разделители тысяч)

    Используем copy() из модуля copy, т.к. объекты стилей в openpyxl
    нельзя просто присвоить напрямую между разными книгами (workbook) —
    они привязаны к внутренней таблице стилей конкретной книги.
    """
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)          # ← цвет заливки ячейки
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def copy_row(ws_src: Worksheet, src_row: int, ws_dst: Worksheet, dst_row: int) -> None:
    """Копирует всю строку (все колонки) с сохранением стилей каждой ячейки."""
    for col_idx in range(1, ws_src.max_column + 1):
        src_cell = ws_src.cell(row=src_row, column=col_idx)
        dst_cell = ws_dst.cell(row=dst_row, column=col_idx)
        copy_cell(src_cell, dst_cell)

    # ─── Сохраняем высоту строки, если она задавалась вручную ────────────────
    src_dim = ws_src.row_dimensions.get(src_row)
    if src_dim and src_dim.height:
        ws_dst.row_dimensions[dst_row].height = src_dim.height


def copy_column_widths(ws_src: Worksheet, ws_dst: Worksheet) -> None:
    """Копирует ширину колонок — иначе итоговый файл будет со сжатыми столбцами."""
    for col_idx in range(1, ws_src.max_column + 1):
        col_letter = get_column_letter(col_idx)
        src_dim = ws_src.column_dimensions.get(col_letter)
        if src_dim and src_dim.width:
            ws_dst.column_dimensions[col_letter].width = src_dim.width


def copy_merged_cells_in_range(
    ws_src: Worksheet, ws_dst: Worksheet, row_start: int, row_end: int, row_offset: int
) -> None:
    """
    Копирует объединённые ячейки (merged cells), если они попадают
    в копируемый диапазон строк. row_offset — сдвиг для позиции в новом файле.

    Пример: если в исходнике объединены A1:C1 (шапка), и мы копируем
    строки 1..1, то в результате тоже будет объединение A1:C1.
    """
    for merged_range in ws_src.merged_cells.ranges:
        if merged_range.min_row >= row_start and merged_range.max_row <= row_end:
            new_min_row = merged_range.min_row - row_start + row_offset
            new_max_row = merged_range.max_row - row_start + row_offset
            try:
                ws_dst.merge_cells(
                    start_row=new_min_row, start_column=merged_range.min_col,
                    end_row=new_max_row, end_column=merged_range.max_col,
                )
            except Exception:
                pass  # пропускаем некорректные диапазоны, не критично


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Сохраняет Workbook в bytes через BytesIO — без записи на диск."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def worksheet_preview_df(ws: Worksheet, max_rows: int = 10) -> pd.DataFrame:
    """
    Строит DataFrame ТОЛЬКО для визуального превью в интерфейсе Streamlit.
    ⚡ Не используется для сохранения файлов — только для st.dataframe().
    """
    header = [cell.value for cell in ws[HEADER_ROW]]
    rows = []
    for row_idx in range(HEADER_ROW + 1, min(ws.max_row, HEADER_ROW + max_rows) + 1):
        rows.append([cell.value for cell in ws[row_idx]])
    return pd.DataFrame(rows, columns=header)


# ════════════════════════════════════════════════════════════════════════════
# БЛОК 1: РАЗДЕЛЕНИЕ
# ════════════════════════════════════════════════════════════════════════════

def load_workbook_safe(uploaded_file) -> tuple[Workbook | None, str | None]:
    """
    Безопасно загружает Excel-файл через openpyxl (сохраняет стили).

    Returns:
        (Workbook, None) при успехе
        (None, error_message) при ошибке
    """
    try:
        wb = load_workbook(uploaded_file, data_only=False)
    except Exception as e:
        return None, f"Не удалось прочитать файл: {e}"

    ws = wb.active
    if ws.max_row <= HEADER_ROW:
        return None, "Файл пуст (нет строк данных, кроме заголовка)"

    return wb, None


def get_data_row_count(ws: Worksheet) -> int:
    """Количество строк данных (без учёта заголовка)."""
    return ws.max_row - HEADER_ROW


def split_workbook(wb: Workbook, chunk_size: int) -> list[dict]:
    """
    Делит книгу на части по chunk_size строк ДАННЫХ (заголовок дублируется
    в каждую часть). Стили, ширина колонок и объединённые ячейки заголовка
    сохраняются в каждой части.

    Returns:
        [
            {"number": 1, "start": 1, "end": 500, "count": 500,
             "wb": <Workbook>, "excel_bytes": b"..."},
            ...
        ]
    """
    ws_src = wb.active
    total_data_rows = get_data_row_count(ws_src)
    chunks = []

    for i, start in enumerate(range(0, total_data_rows, chunk_size), start=1):
        end = min(start + chunk_size, total_data_rows)

        wb_chunk = Workbook()
        ws_chunk = wb_chunk.active
        ws_chunk.title = ws_src.title or f"Часть {i}"

        # ─── Копируем заголовок (строка 1) со стилями ────────────────────────
        copy_row(ws_src, HEADER_ROW, ws_chunk, 1)
        copy_merged_cells_in_range(
            ws_src, ws_chunk, HEADER_ROW, HEADER_ROW, row_offset=1
        )

        # ─── Копируем строки данных этого диапазона ──────────────────────────
        dst_row = 2
        for src_row in range(HEADER_ROW + 1 + start, HEADER_ROW + 1 + end):
            copy_row(ws_src, src_row, ws_chunk, dst_row)
            dst_row += 1

        # ─── Копируем ширину колонок ──────────────────────────────────────────
        copy_column_widths(ws_src, ws_chunk)

        chunks.append({
            "number": i,
            "start": start + 1,
            "end": end,
            "count": end - start,
            "wb": wb_chunk,
            "excel_bytes": workbook_to_bytes(wb_chunk),
        })

    return chunks


# ════════════════════════════════════════════════════════════════════════════
# БЛОК 2: РАБОТА С ZIP-АРХИВАМИ
# ════════════════════════════════════════════════════════════════════════════

def extract_excel_from_zip(zip_bytes: bytes, zip_filename: str) -> list[dict]:
    """
    Распаковывает ZIP в памяти, ищет Excel-файл среди любого количества
    файлов в архиве (остальные игнорируются по договорённости).

    Returns:
        [{"filename": str, "source_zip": str, "wb": Workbook|None, "error": str|None}]
    """
    results = []
    zip_buffer = io.BytesIO(zip_bytes)

    try:
        with zipfile.ZipFile(zip_buffer) as zf:
            infos = zf.infolist()

            total_uncompressed = sum(i.file_size for i in infos)
            if total_uncompressed > MAX_UNCOMPRESSED_SIZE_BYTES:
                results.append({
                    "filename": zip_filename, "source_zip": zip_filename,
                    "wb": None,
                    "error": (
                        f"Архив слишком большой в распакованном виде "
                        f"({total_uncompressed / 1024 / 1024:.1f} MB)."
                    ),
                })
                return results

            excel_infos = [
                i for i in infos
                if i.filename.lower().endswith(ALLOWED_EXCEL_EXTENSIONS)
                and not i.is_dir()
                and not i.filename.startswith("__MACOSX")
            ]

            if not excel_infos:
                results.append({
                    "filename": zip_filename, "source_zip": zip_filename,
                    "wb": None,
                    "error": "В архиве не найден Excel-файл (.xlsx/.xls)",
                })
                return results

            for info in excel_infos:
                try:
                    with zf.open(info) as f:
                        content = f.read()

                    wb = load_workbook(io.BytesIO(content), data_only=False)
                    ws = wb.active

                    if ws.max_row <= HEADER_ROW:
                        results.append({
                            "filename": info.filename, "source_zip": zip_filename,
                            "wb": None, "error": "Файл пуст (нет данных)",
                        })
                        continue

                    results.append({
                        "filename": info.filename, "source_zip": zip_filename,
                        "wb": wb, "error": None,
                    })

                except Exception as e:
                    results.append({
                        "filename": info.filename, "source_zip": zip_filename,
                        "wb": None, "error": f"Ошибка чтения файла: {e}",
                    })

    except zipfile.BadZipFile:
        results.append({
            "filename": zip_filename, "source_zip": zip_filename,
            "wb": None, "error": "Повреждённый ZIP-архив",
        })
    except Exception as e:
        results.append({
            "filename": zip_filename, "source_zip": zip_filename,
            "wb": None, "error": f"Неизвестная ошибка: {e}",
        })

    return results


# ════════════════════════════════════════════════════════════════════════════
# БЛОК 3: ОБЪЕДИНЕНИЕ С СОХРАНЕНИЕМ ФОРМАТИРОВАНИЯ
# ════════════════════════════════════════════════════════════════════════════

def merge_workbooks(entries: list[dict]) -> tuple[Workbook, dict]:
    """
    Объединяет несколько книг в одну, построчно копируя стили.

    Логика:
        - Колонки одинаковые во всех файлах
        - Заголовок берётся из первого файла (со стилями)
        - Строки данных всех файлов добавляются друг за другом
        - Дубликаты НЕ удаляются
    """

    valid_entries = [e for e in entries if e["wb"] is not None]

    if not valid_entries:
        raise ValueError("Нет валидных данных для объединения")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Итоговый отчёт"

    # ─── Заголовок и ширина колонок — из первого файла ───────────────────────
    first_ws = valid_entries[0]["wb"].active
    copy_row(first_ws, HEADER_ROW, ws_out, 1)
    copy_merged_cells_in_range(first_ws, ws_out, HEADER_ROW, HEADER_ROW, row_offset=1)
    copy_column_widths(first_ws, ws_out)

    dst_row = 2
    total_rows_added = 0

    for entry in valid_entries:
        ws_src = entry["wb"].active

        for src_row in range(HEADER_ROW + 1, ws_src.max_row + 1):
            copy_row(ws_src, src_row, ws_out, dst_row)
            dst_row += 1
            total_rows_added += 1

    stats = {
        "total_source_files": len(valid_entries),
        "invalid_files_count": len(entries) - len(valid_entries),
        "rows_total": total_rows_added,
    }

    return wb_out, stats


# ════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Excel Split & Merge", page_icon="🗂️", layout="wide")

st.markdown("""
<style>
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


def init_state():
    defaults = {
        "original_wb": None,
        "original_filename": None,
        "chunks": [],
        "extracted_entries": [],
        "merged_wb": None,
        "merged_stats": None,
        "merged_bytes": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


st.title("🗂️ Excel: Разделение и Объединение")
st.caption("Форматирование (цвета ячеек, шрифты, границы) сохраняется на всех этапах")

tab_split, tab_merge = st.tabs(["✂️ Разделение", "🔗 Объединение"])


# ────────────────────────────────────────────────────────────────────────────
# TAB 1: РАЗДЕЛЕНИЕ
# ────────────────────────────────────────────────────────────────────────────

with tab_split:
    st.markdown("### 1️⃣ Загрузите исходный Excel-файл")

    uploaded_original = st.file_uploader(
        "Excel-файл (.xlsx)", type=["xlsx", "xls"], key="original_uploader",
    )

    if uploaded_original is not None:
        if st.session_state.original_filename != uploaded_original.name:
            wb, error = load_workbook_safe(uploaded_original)

            if error:
                st.error(f"❌ {error}")
                st.session_state.original_wb = None
            else:
                st.session_state.original_wb = wb
                st.session_state.original_filename = uploaded_original.name
                st.session_state.chunks = []
                ws = wb.active
                st.success(
                    f"✅ Загружено {get_data_row_count(ws)} строк данных, "
                    f"{ws.max_column} колонок"
                )

    wb = st.session_state.original_wb

    if wb is not None:
        ws = wb.active

        st.markdown("#### Предпросмотр данных")
        st.dataframe(worksheet_preview_df(ws), use_container_width=True)
        st.caption(
            "⚡ Цвета и форматирование не видны в этом превью "
            "(ограничение Streamlit-таблицы), но сохраняются в скачиваемых файлах"
        )

        total_rows = get_data_row_count(ws)
        half_limit = total_rows // 2

        st.markdown("### 2️⃣ Настройте размер части")

        if half_limit < 10:
            st.warning(
                f"⚠️ В документе слишком мало строк ({total_rows}) "
                f"для разбивки минимум по 10 строк в части."
            )
            st.stop()

        col_input, col_info = st.columns([1, 2])

        with col_input:
            chunk_size = st.number_input(
                "Строк в одной части", min_value=10,
                value=min(500, max(10, half_limit)), step=10,
            )

        is_invalid_size = chunk_size > half_limit

        with col_info:
            st.markdown("<br>", unsafe_allow_html=True)
            if is_invalid_size:
                st.error(
                    f"⚠️ Размер части не может быть больше половины документа "
                    f"({half_limit} строк)."
                )
            else:
                estimated_parts = -(-total_rows // chunk_size)
                st.info(f"Будет создано частей: **{estimated_parts}**")

        st.markdown("### 3️⃣ Разбейте на части")

        if st.button("✂️ Разбить на части", type="primary", disabled=is_invalid_size):
            with st.spinner("Разбиваем документ (сохраняем форматирование)..."):
                st.session_state.chunks = split_workbook(wb, chunk_size)
            st.success(f"✅ Создано {len(st.session_state.chunks)} частей")

        if st.session_state.chunks:
            st.markdown("### 4️⃣ Скачайте части")

            base_name = st.session_state.original_filename.rsplit(".", 1)[0]
            cols_per_row = 4
            chunks = st.session_state.chunks

            for row_start in range(0, len(chunks), cols_per_row):
                row = chunks[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)

                for col, chunk in zip(cols, row):
                    with col:
                        st.markdown(
                            f"**Часть {chunk['number']}**  \n"
                            f"Строки {chunk['start']}–{chunk['end']}  \n"
                            f"({chunk['count']} шт.)"
                        )
                        st.download_button(
                            label="⬇️ Скачать",
                            data=chunk["excel_bytes"],
                            file_name=f"{base_name}_part_{chunk['number']}.xlsx",
                            mime=(
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            ),
                            key=f"dl_split_{chunk['number']}",
                            use_container_width=True,
                        )
    else:
        st.info("👆 Загрузите Excel-файл, чтобы начать")


# ────────────────────────────────────────────────────────────────────────────
# TAB 2: ОБЪЕДИНЕНИЕ
# ────────────────────────────────────────────────────────────────────────────

with tab_merge:
    st.markdown("### 1️⃣ Загрузите ZIP-архивы с результатами")
    st.caption("Каждый архив содержит ровно 1 Excel-файл — остальные файлы игнорируются")

    uploaded_zips = st.file_uploader(
        "ZIP-архивы", type=["zip"], accept_multiple_files=True, key="zip_uploader",
    )

    if uploaded_zips:
        if st.button("📦 Обработать архивы", type="primary"):
            all_entries = []
            progress_bar = st.progress(0)
            status_text = st.empty()

            for idx, zip_file in enumerate(uploaded_zips):
                status_text.text(f"Обрабатываем: {zip_file.name}")
                zip_bytes = zip_file.read()
                entries = extract_excel_from_zip(zip_bytes, zip_file.name)
                all_entries.extend(entries)
                progress_bar.progress((idx + 1) / len(uploaded_zips))

            status_text.empty()
            progress_bar.empty()

            st.session_state.extracted_entries = all_entries
            st.session_state.merged_wb = None
            st.session_state.merged_bytes = None

    entries = st.session_state.extracted_entries

    if entries:
        st.markdown("### 2️⃣ Найденные файлы")

        valid_count = sum(1 for e in entries if e["wb"] is not None)
        error_count = len(entries) - valid_count

        c1, c2, c3 = st.columns(3)
        c1.metric("Всего архивов", len(entries))
        c2.metric("✅ Валидных", valid_count)
        c3.metric("❌ С ошибками", error_count)

        for e in entries:
            if e["wb"] is not None:
                ws_e = e["wb"].active
                st.success(
                    f"✅ `{e['filename']}` (из {e['source_zip']}) — "
                    f"{get_data_row_count(ws_e)} строк"
                )
            else:
                st.error(f"❌ `{e['filename']}` (из {e['source_zip']}) — {e['error']}")

        if valid_count > 0:
            st.markdown("### 3️⃣ Объедините данные")

            if st.button("🔗 Объединить все файлы", type="primary"):
                with st.spinner("Объединяем (сохраняем форматирование)..."):
                    try:
                        merged_wb, stats = merge_workbooks(entries)
                        st.session_state.merged_wb = merged_wb
                        st.session_state.merged_stats = stats
                        st.session_state.merged_bytes = workbook_to_bytes(merged_wb)
                    except ValueError as e:
                        st.error(f"❌ {e}")

            if st.session_state.merged_wb is not None:
                stats = st.session_state.merged_stats

                st.markdown("#### 📊 Статистика объединения")
                s1, s2, s3 = st.columns(3)
                s1.metric("Файлов объединено", stats["total_source_files"])
                s2.metric("Строк всего", stats["rows_total"])
                s3.metric("Файлов с ошибками", stats["invalid_files_count"])

                st.markdown("#### Предпросмотр итогового файла")
                st.dataframe(
                    worksheet_preview_df(st.session_state.merged_wb.active, max_rows=20),
                    use_container_width=True,
                )
                st.caption("⚡ Цвета ячеек сохранены в скачиваемом файле")

                st.markdown("### 4️⃣ Скачайте итоговый файл")

                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    label="⬇️ Скачать итоговый Excel",
                    data=st.session_state.merged_bytes,
                    file_name=f"merged_report_{timestamp}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),
                    type="primary",
                    use_container_width=True,
                )
        else:
            st.warning("Нет ни одного валидного Excel-файла для объединения.")
    else:
        st.info("👆 Загрузите ZIP-архивы, чтобы начать объединение")